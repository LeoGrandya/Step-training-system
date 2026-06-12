# -*- coding: utf-8 -*-
"""
配置加载器：合并 user_params.py + 亮灯间隔 Excel + 受试者信息 Excel
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Optional, Tuple

from openpyxl import load_workbook

from user_params import USER_PARAMS


def _resolve_path(raw: str, project_root: Path) -> Path:
    """将相对路径转为绝对路径。"""
    p = Path(raw)
    if p.is_absolute():
        return p
    return (project_root / p).resolve()


def load_config(project_root: Path, person_name: str, video_id: str, group_name: str) -> Dict[str, Any]:
    """
    加载并合并所有配置源。

    Parameters
    ----------
    project_root : Path
        项目根目录。
    person_name : str
        运动员姓名。
    video_id : str
        视频编号，如 "5-1"。
    group_name : str
        组名，如 "adult_video1"。

    Returns
    -------
    dict
        合并后的完整参数字典。
    """
    params = USER_PARAMS.copy()

    # ---------- 解析路径 ----------
    input_root = _resolve_path(params.get("input_root", "./input"), project_root)
    output_root = _resolve_path(params.get("output_root", "./output"), project_root)
    subject_xlsx = _resolve_path(params.get("subject_info_xlsx", "./input/受试者基本信息.xlsx"), project_root)
    light_xlsx = _resolve_path(params.get("light_interval_xlsx", "./input/亮灯时间间隔配置.xlsx"), project_root)

    params["_input_root"] = input_root
    params["_output_root"] = output_root

    # ---------- 受试者信息 ----------
    subject_info = _load_subject_info(subject_xlsx, person_name, params)
    if subject_info.get("height_m") is None or subject_info.get("weight_kg") is None:
        raise ValueError(
            f"在 {subject_xlsx} 中未找到运动员 '{person_name}' 的身高/体重信息。"
            f"请在 Excel 中补充该运动员的数据后再运行。"
        )
    params["body"] = {
        "height_m": subject_info["height_m"],
        "weight_kg": subject_info["weight_kg"],
    }
    params["handedness"] = subject_info.get("handedness", "right")

    # ---------- 亮灯间隔 ----------
    params["light_interval_ms"] = _load_light_interval(light_xlsx, person_name, video_id, group_name)

    return params


def _load_subject_info(xlsx_path: Path, person_name: str, params: Dict[str, Any]) -> Dict[str, Any]:
    """从受试者基本信息.xlsx 读取身高、体重、优势手。"""
    col_map = params.get("subject_excel_columns", {})
    col_name = col_map.get("name", 0)
    col_height = col_map.get("height_cm", 3)
    col_weight = col_map.get("weight_kg", 4)
    col_hand = col_map.get("handedness", 7)

    result: Dict[str, Any] = {}

    if not xlsx_path.exists():
        raise FileNotFoundError(f"受试者信息表不存在：{xlsx_path}")

    try:
        wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[col_name] is None:
                continue
            name = str(row[col_name]).strip()
            if name == person_name:
                if row[col_height] is not None:
                    result["height_m"] = float(row[col_height]) / 100.0
                if row[col_weight] is not None:
                    result["weight_kg"] = float(row[col_weight])
                if row[col_hand] is not None:
                    hand_raw = str(row[col_hand]).strip()
                    # 归一化: 右/右手 → right, 左 → left
                    if "左" in hand_raw:
                        result["handedness"] = "left"
                    else:
                        result["handedness"] = "right"
                break
        wb.close()
    except Exception:
        pass

    return result


def _load_light_interval(
    xlsx_path: Path, person_name: str, video_id: str, group_name: str
) -> Optional[float]:
    """
    从亮灯时间间隔配置.xlsx 读取亮灯间隔。

    Parameters
    ----------
    person_name : str
        运动员姓名。
    video_id : str
        视频编号，如 "5-1"。
    group_name : str
        组名，同时也是 Excel 的 sheet 名，如 "adult_video1"。

    Returns
    -------
    float or None
        亮灯间隔（秒），None 表示此视频无亮灯配置（4/5/6/7 步伐不需要）。
    """
    if not xlsx_path.exists():
        return None

    try:
        wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)

        # 先尝试按 group_name 找 sheet，找不到则尝试所有 sheet
        sheet = None
        if group_name in wb.sheetnames:
            sheet = wb[group_name]
        else:
            for sn in wb.sheetnames:
                sheet = wb[sn]
                break

        if sheet is None:
            wb.close()
            return None

        # 找表头行：video_id 所在的列索引
        header = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        video_col_idx = None
        for i, val in enumerate(header):
            if val is not None and str(val).strip() == video_id:
                video_col_idx = i
                break

        if video_col_idx is None:
            wb.close()
            return None

        # 找运动员行
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and str(row[0]).strip() == person_name:
                val = row[video_col_idx]
                wb.close()
                if val is not None:
                    # Excel 中的值单位是毫秒，转为秒
                    return float(val) / 1000.0
                return None

        wb.close()
        return None
    except Exception:
        return None


def parse_pace_id(video_dir_name: str) -> int:
    """从目录名解析步伐编号。支持 "1-1" 和 "跨步-1" 两种格式。"""
    parts = video_dir_name.strip().split("-")
    first = parts[0]
    try:
        return int(first)
    except ValueError:
        pass
    if first in PACE_NAME_TO_ID:
        return PACE_NAME_TO_ID[first]
    raise ValueError(f"无法解析步伐编号: {video_dir_name}")


PACE_NAMES: Dict[int, str] = {
    1: "跨步",
    2: "并步",
    3: "撤步",
    4: "两点跑位",
    5: "三点跑位",
    6: "四点跑位",
    7: "推侧扑",
    8: "全台摆速",
}

# 中文名 → 编号（支持中文目录名）
PACE_NAME_TO_ID: Dict[str, int] = {v: k for k, v in PACE_NAMES.items()}
